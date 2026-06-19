#!/usr/bin/env python3
"""Append two SIT testcase rows (chained titlefetch -> directposting) to sit_testcases.xlsx

Usage: python scripts/add_sit_testcases.py
"""
from openpyxl import load_workbook
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
TESTCASE_FILE = ROOT / "testcases" / "sit_testcases.xlsx"
SHEET_NAME = "SIT"


def main():
    wb = load_workbook(TESTCASE_FILE)
    if SHEET_NAME in wb.sheetnames:
        sheet = wb[SHEET_NAME]
    else:
        sheet = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

    def make_row(values_dict):
        return [values_dict.get(h, "") for h in headers]

    # Positive chained case: titlefetch success -> directposting
    row_pos = {
        "TC_ID": "TC176",
        "Description": "Chained: TitleFetch success -> DirectPosting executed",
        "Bank": "bank1",
        "API_Name": "Chained-TitleFetch-DirectPosting",
        "Request_File": "tc176_title_fetch.json",
        "Expected_Result": "00",
        "Validation_Type": "response",
        "Execute (Y/N)": "Y",
        "Flow_Config": "Chained-TitleFetch-DirectPosting",
        "Test_Type": "positive",
    }

    # Negative chained case: titlefetch failure -> no directposting
    row_neg = {
        "TC_ID": "TC177",
        "Description": "Chained: TitleFetch failure -> DirectPosting NOT called",
        "Bank": "bank1",
        "API_Name": "Chained-TitleFetch-DirectPosting",
        "Request_File": "tc178_title_fetch_negative.json",
        "Expected_Result": "101",
        "Validation_Type": "response",
        "Execute (Y/N)": "Y",
        "Flow_Config": "Chained-TitleFetch-DirectPosting",
        "Test_Type": "negative",
    }

    # Append rows
    sheet.append(make_row(row_pos))
    sheet.append(make_row(row_neg))

    wb.save(TESTCASE_FILE)
    print(f"Appended TC176 and TC177 to {TESTCASE_FILE}")


if __name__ == "__main__":
    main()
