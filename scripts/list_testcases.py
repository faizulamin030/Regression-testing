from openpyxl import load_workbook
from pathlib import Path

wb = load_workbook(Path('testcases') / 'sit_testcases.xlsx', data_only=True)
sheet = wb['SIT'] if 'SIT' in wb.sheetnames else wb.active
headers = [str(c.value).strip() if c.value else '' for c in sheet[1]]
for row in sheet.iter_rows(min_row=2, values_only=True):
    tc_id = str(row[headers.index('TC_ID')]) if 'TC_ID' in headers else ''
    if tc_id in ('TC176','TC177','TC178'):
        d = {headers[i]: row[i] for i in range(len(headers))}
        print(tc_id, '->', d.get('Request_File'), 'Flow_Config->', d.get('Flow_Config'))
