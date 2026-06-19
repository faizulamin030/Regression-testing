import copy
import json
import os
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import yaml
from openpyxl import load_workbook

from utils.api_client import ApiClient
from utils.data_generator import DataGenerator
from utils.logger import LoggerFactory
from utils.service_manager import ServiceManager
from utils.template_engine import TemplateEngine
from validators.response_validator import ResponseValidator


@dataclass
class TestCase:
    """Legacy test case without service awareness."""

    tc_id: str
    description: str
    bank: str
    api_name: str
    request_file: str
    expected_result: str
    validation_type: str
    execute: str
    flow_config: str = ""
    validation_rules: str = ""


@dataclass
class ServiceTestCase:
    """Service-aware test case with service context."""

    tc_id: str
    service: str
    description: str
    bank: str
    api_name: str
    request_file: str
    expected_result: str
    validation_type: str
    execute: str
    flow_config: str = ""
    validation_rules: str = ""
    test_type: str = "positive"


class SITRunner:
    def __init__(self, root_dir: Path) -> None:
        self.root_dir = root_dir
        self.config_dir = root_dir / "config"
        self.requests_dir = root_dir / "requests"
        self.testcases_dir = root_dir / "testcases"
        self.testcase_sources = [
            {
                "service": "cas",
                "file": self.testcases_dir / "sit_testcases.xlsx",
                "sheet": "SIT",
            },
            {
                "service": "acquiring",
                "file": self.testcases_dir / "sit_testcases_acquiring.xlsx",
                "sheet": "ACQUIRING",
            },
            {
                "service": "merchant_cas",
                "file": self.testcases_dir / "sit_testcases_merchant_cas.xlsx",
                "sheet": "MERCHANT_CAS",
            },
        ]
        self.results_dir = root_dir / "results"
        self.reports_dir = root_dir / "reports"
        self.logs_dir = root_dir / "logs"

        for directory in [self.results_dir, self.reports_dir, self.logs_dir]:
            directory.mkdir(parents=True, exist_ok=True)
        self.rrn_registry_file = self.results_dir / "rrn_registry.json"
        DataGenerator.configure_rrn_registry(self.rrn_registry_file)

        logger_factory = LoggerFactory(self.logs_dir)
        logger_factory.configure()
        self.framework_logger = logger_factory.get("framework")
        self.request_logger = logger_factory.get("request")
        self.response_logger = logger_factory.get("response")
        self.validation_logger = logger_factory.get("validation")
        self.error_logger = logger_factory.get("error")

        self.server_config = self._load_yaml(self.config_dir / "servers.yaml")
        self.api_config = self._load_yaml(self.config_dir / "apis.yaml")
        self.reporting_config = self._load_reporting_config()

        # Load user-defined test data variables (config/test_data.yaml)
        test_data_file = self.config_dir / "test_data.yaml"
        raw_test_data = self._load_yaml(test_data_file) if test_data_file.exists() else {}
        self.test_data_tokens: Dict[str, str] = {
            str(k).upper(): str(v) for k, v in raw_test_data.items() if v is not None
        }

        # Initialize ServiceManager with merged service configuration from
        # servers.yaml (infra/base_url/enabled) and apis.yaml (auth/apis/token_routing).
        services_config = self._build_merged_services_config(
            self.server_config.get("services", {}),
            self.api_config.get("services", {}),
        )
        self.service_manager = ServiceManager(services_config)

        # Apply environment-based service filtering
        self._apply_env_service_filter()

        # Log enabled services
        enabled_services = self.service_manager.get_services_by_status(enabled_only=True)
        self.framework_logger.info("Enabled services: %s", ", ".join(enabled_services))

        # Service-scoped API configuration
        self.service_api_config = self.api_config.get("services", {})

        # Service-scoped token contexts
        self.service_token_context: Dict[str, Dict[str, str]] = {
            service_id: {} for service_id in self.service_manager.get_services_by_status(enabled_only=False)
        }

        defaults = self.api_config.get("defaults", {})
        self.api_client = ApiClient(
            timeout_seconds=int(defaults.get("timeout_seconds", 30)),
            retries=int(defaults.get("retries", 2)),
            retry_backoff_factor=float(defaults.get("retry_backoff_factor", 0.5)),
            retry_status_codes=list(defaults.get("retry_status_codes", [500, 502, 503, 504])),
            request_logger=self.request_logger,
            response_logger=self.response_logger,
            error_logger=self.error_logger,
        )
        self.validator = ResponseValidator(self.validation_logger)
        self.runtime_context: Dict[str, Any] = {}
        # Legacy token context for backward compatibility
        self.token_context: Dict[str, str] = {}

    def _build_merged_services_config(
        self,
        server_services: Dict[str, Any],
        api_services: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Merge per-service settings from servers and APIs config files."""
        merged: Dict[str, Any] = {}

        service_ids = set(server_services.keys()) | set(api_services.keys())
        for service_id in service_ids:
            merged_entry: Dict[str, Any] = {}

            server_entry = server_services.get(service_id, {})
            if isinstance(server_entry, dict):
                merged_entry.update(copy.deepcopy(server_entry))

            api_entry = api_services.get(service_id, {})
            if isinstance(api_entry, dict):
                if "display_name" in api_entry:
                    merged_entry["display_name"] = api_entry["display_name"]
                for key in ["auth", "apis", "token_routing"]:
                    if key in api_entry:
                        merged_entry[key] = copy.deepcopy(api_entry[key])

            merged_entry.setdefault("display_name", service_id)
            merged_entry.setdefault("enabled", True)
            merged_entry.setdefault("base_url", "")
            merged_entry.setdefault("auth", {})
            merged_entry.setdefault("apis", {})
            merged_entry.setdefault("token_routing", {})
            merged[service_id] = merged_entry

        return merged

    def _load_yaml(self, file_path: Path) -> Dict[str, Any]:
        with file_path.open("r", encoding="utf-8") as handle:
            return yaml.safe_load(handle) or {}

    def _load_reporting_config(self) -> Dict[str, Any]:
        defaults = {
            "html_report": {
                "hide_sensitive_data": True,
                "show_failure_analysis": True,
                "show_details": True,
            }
        }
        reporting_file = self.config_dir / "reporting.yaml"
        if not reporting_file.exists():
            return defaults

        loaded = self._load_yaml(reporting_file)
        html_report = loaded.get("html_report", {}) if isinstance(loaded, dict) else {}
        defaults["html_report"].update(
            {
                key: value
                for key, value in html_report.items()
                if key in defaults["html_report"]
            }
        )
        return defaults

    def _reporting_bool(self, path: str, default: bool) -> bool:
        current: Any = self.reporting_config
        for part in path.split("."):
            if not isinstance(current, dict) or part not in current:
                return default
            current = current[part]

        if isinstance(current, bool):
            return current
        if isinstance(current, str):
            return current.strip().lower() in {"1", "true", "yes", "y", "on"}
        return bool(current)

    def _apply_env_service_filter(self) -> None:
        """Apply environment-based service filtering (enable/disable)."""
        enabled_services_str = os.getenv("ENABLED_SERVICES", "").strip()
        disabled_services_str = os.getenv("DISABLED_SERVICES", "").strip()

        if enabled_services_str:
            enabled_list = [s.strip().lower() for s in enabled_services_str.split(",") if s.strip()]
            self.service_manager.apply_env_filter(enabled_services=enabled_list)

        if disabled_services_str:
            disabled_list = [s.strip().lower() for s in disabled_services_str.split(",") if s.strip()]
            self.service_manager.apply_env_filter(disabled_services=disabled_list)

    def load_testcases(self) -> List[ServiceTestCase]:
        """Load testcases from service-specific files/sheets when available."""
        required = [
            "TC_ID",
            "Description",
            "Bank",
            "API_Name",
            "Request_File",
            "Expected_Result",
            "Validation_Type",
            "Execute (Y/N)",
        ]

        testcases: List[ServiceTestCase] = []

        for source in self.testcase_sources:
            source_service = str(source["service"])
            source_file = Path(source["file"])
            source_sheet = str(source["sheet"])

            if not source_file.exists():
                self.framework_logger.info(
                    "Testcase source not found, skipping: service=%s file=%s",
                    source_service,
                    source_file,
                )
                continue

            workbook = load_workbook(filename=source_file)
            if source_sheet in workbook.sheetnames:
                sheet = workbook[source_sheet]
            else:
                sheet = workbook.active
                self.framework_logger.warning(
                    "Sheet '%s' not found in %s. Using active sheet '%s'.",
                    source_sheet,
                    source_file.name,
                    sheet.title,
                )

            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
            header_index = {header: idx for idx, header in enumerate(headers)}

            missing = [column for column in required if column not in header_index]
            if missing:
                raise ValueError(
                    f"Missing required columns in Excel '{source_file.name}' sheet '{sheet.title}': {missing}"
                )

            loaded_for_source = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue

                service = self._resolve_service_for_row(row, header_index, default_service=source_service)

                if not self.service_manager.is_service_enabled(service):
                    self.framework_logger.debug(
                        "Skipped test case (disabled service): tc_id=%s, service=%s",
                        str(row[header_index["TC_ID"]]).strip(),
                        service,
                    )
                    continue

                testcase = ServiceTestCase(
                    tc_id=str(row[header_index["TC_ID"]]).strip(),
                    service=service,
                    description=str(row[header_index.get("Description", -1)] or "").strip(),
                    bank=str(row[header_index.get("Bank", -1)] or "").strip(),
                    api_name=str(row[header_index.get("API_Name", -1)] or "").strip(),
                    request_file=str(row[header_index.get("Request_File", -1)] or "").strip(),
                    expected_result=str(row[header_index.get("Expected_Result", -1)] or "").strip(),
                    validation_type=str(row[header_index.get("Validation_Type", -1)] or "response").strip(),
                    execute=str(row[header_index["Execute (Y/N)"]]).strip().upper(),
                    flow_config=str(row[header_index.get("Flow_Config", -1)] or "").strip()
                    if "Flow_Config" in header_index
                    else "",
                    validation_rules=str(row[header_index.get("Validation_Rules", -1)] or "").strip()
                    if "Validation_Rules" in header_index
                    else "",
                    test_type=str(row[header_index.get("Test_Type", -1)] or "positive").strip().lower()
                    if "Test_Type" in header_index
                    else "positive",
                )
                if testcase.execute == "Y":
                    testcases.append(testcase)
                    loaded_for_source += 1

            self.framework_logger.info(
                "Loaded %d testcase(s) from service=%s, file=%s, sheet=%s",
                loaded_for_source,
                source_service,
                source_file.name,
                sheet.title,
            )

        return testcases

    def _get_token_for_api(self, api_name: str, service: str = "cas") -> str:
        """Get token for API, with service context awareness."""
        token_routing = self.service_manager.get_service_token_routing(service)
        if api_name in token_routing:
            channel_or_key = token_routing[api_name]
            return self.service_token_context.get(service, {}).get(channel_or_key, "")
        return self.service_token_context.get(service, {}).get("default", "")

    def run(self) -> None:
        started_at = datetime.now()
        self.framework_logger.info("Execution started at %s", started_at.isoformat())

        max_workers = int(os.getenv("MAX_WORKERS", "4"))
        testcases = self.load_testcases()
        
        # Group testcases by service
        testcases_by_service = {}
        for tc in testcases:
            if tc.service not in testcases_by_service:
                testcases_by_service[tc.service] = []
            testcases_by_service[tc.service].append(tc)
        
        # Log test case counts per service
        for service, tcs in testcases_by_service.items():
            self.framework_logger.info("[%s] %d test cases loaded", service, len(tcs))
        
        results: List[Dict[str, Any]] = []
        lock = threading.Lock()
        
        # Execute per service, handling auth within each service context
        for service in self.service_manager.get_services_by_status(enabled_only=True):
            if service not in testcases_by_service:
                continue
            
            service_testcases = testcases_by_service[service]
            auth_api_names = self._auth_api_names(service)
            auth_testcases = [tc for tc in service_testcases if tc.api_name in auth_api_names]
            non_auth_testcases = [tc for tc in service_testcases if tc.api_name not in auth_api_names]
            independent_testcases = [
                tc for tc in non_auth_testcases if not self._is_sequential_testcase(tc.tc_id)
            ]
            sequential_testcases = [
                tc for tc in non_auth_testcases if self._is_sequential_testcase(tc.tc_id)
            ]
            sequential_testcases.sort(key=lambda tc: self._tc_order_key(tc.tc_id))
            
            # Execute auth test cases sequentially
            for auth_tc in auth_testcases:
                auth_result = self.execute_testcase(auth_tc)
                results.append(auth_result)
                if auth_result.get("status") != "PASSED":
                    self.framework_logger.warning(
                        "[%s] Auth testcase failed. tc_id=%s. Subsequent APIs may run without bearer token.",
                        service,
                        auth_tc.tc_id,
                    )
            
            # Execute independent non-auth test cases in parallel.
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = [executor.submit(self.execute_testcase, testcase) for testcase in independent_testcases]
                for future in as_completed(futures):
                    result = future.result()
                    with lock:
                        results.append(result)

            # Execute dependent test cases strictly in sequence.
            for sequential_tc in sequential_testcases:
                result = self.execute_testcase(sequential_tc)
                results.append(result)

        # Sort results: TC_AUTH_IB first, then remaining by service and TC_ID
        results.sort(key=lambda x: (
            x.get("service", "unknown"),
            x.get("tc_id", "") != "TC_AUTH_IB",
            x.get("tc_id", "")
        ))

        finished_at = datetime.now()
        summary = self._build_summary(results)
        run_tag = finished_at.strftime("%Y%m%d_%H%M%S")
        self._write_results_json(run_tag, results, summary)
        self._write_results_csv(run_tag, results)
        self._write_html_report(run_tag, results, summary, started_at, finished_at)
        self.framework_logger.info("Execution completed. Summary=%s", summary)

    def execute_testcase(self, testcase: ServiceTestCase) -> Dict[str, Any]:
        """Execute a service test case with service context awareness."""
        try:
            bank_base_url = self._get_bank_base_url(testcase.bank, testcase.service)
            api_definition = self._get_api_definition(testcase.api_name, testcase.service)

            if testcase.flow_config:
                return self._execute_flow_testcase(testcase, bank_base_url)

            payload = self._render_payload(testcase.request_file, extra_tokens={}, service=testcase.service)
            url = self._build_url(bank_base_url, api_definition["endpoint"])
            headers = self._build_request_headers(api_definition, testcase.api_name, testcase.service)
            response = self.api_client.call(
                method=str(api_definition["method"]).upper(),
                url=url,
                payload=payload,
                headers=headers,
            )

            self._capture_auth_token_if_configured(testcase, response)

            validation_rules = self._parse_json_object(testcase.validation_rules)
            validation = self.validator.validate(
                validation_type=testcase.validation_type,
                expected_result=testcase.expected_result,
                response=response,
                validation_rules=validation_rules,
            )

            return {
                "tc_id": testcase.tc_id,
                "service": testcase.service,
                "description": testcase.description,
                "bank": testcase.bank,
                "api_name": testcase.api_name,
                "test_type": testcase.test_type,
                "status": "PASSED" if validation["passed"] else "FAILED",
                "expected": testcase.expected_result,
                "actual": validation["actual"],
                "status_code": response.get("status_code"),
                "response_time_ms": response.get("elapsed_ms"),
                "request_payload": payload,
                "request_headers": headers,
                "response_body": response.get("body"),
                "error": "",
            }
        except Exception as exc:
            self.error_logger.exception(
                "TESTCASE_ERROR | service=%s | tc_id=%s | error=%s",
                testcase.service,
                testcase.tc_id,
                str(exc),
            )
            return {
                "tc_id": testcase.tc_id,
                "service": testcase.service,
                "description": testcase.description,
                "bank": testcase.bank,
                "api_name": testcase.api_name,
                "test_type": testcase.test_type,
                "status": "FAILED",
                "expected": testcase.expected_result,
                "actual": str(exc),
                "status_code": None,
                "response_time_ms": 0,
                "request_payload": {},
                "request_headers": {},
                "response_body": {},
                "error": str(exc),
            }

    def _execute_flow_testcase(self, testcase: ServiceTestCase, bank_base_url: str) -> Dict[str, Any]:
        """Execute a flow (multi-step) test case with service context."""
        # Get service-specific flows
        service_flows = self.service_api_config.get(testcase.service, {}).get("flows", {})
        flow = service_flows.get(testcase.flow_config)
        
        if not flow:
            raise ValueError(f"Flow config not found in service '{testcase.service}': {testcase.flow_config}")
        # Support two flow styles:
        # 1) existing polling-based flows (polling key)
        # 2) conditional "on_success" chained flows where a second API is called only if the initial response indicates success

        initial_api_name = flow.get("initial_api", testcase.api_name)
        initial_request_file = flow.get("initial_request_file", testcase.request_file)
        initial_api_definition = self._get_api_definition(initial_api_name, testcase.service)
        initial_payload = self._render_payload(initial_request_file, extra_tokens={}, service=testcase.service)
        initial_url = self._build_url(bank_base_url, initial_api_definition["endpoint"])

        initial_headers = self._build_request_headers(initial_api_definition, initial_api_name, testcase.service)
        initial_response = self.api_client.call(
            method=str(initial_api_definition["method"]).upper(),
            url=initial_url,
            payload=initial_payload,
            headers=initial_headers,
        )

        # If flow uses polling (legacy), keep previous behavior
        polling_definition = flow.get("polling")
        if polling_definition:
            poll_response = self._poll_for_final_status(
                bank_base_url=bank_base_url,
                polling_definition=polling_definition,
                initial_response=initial_response,
                service=testcase.service,
            )

            validation = self.validator.validate(
                validation_type=testcase.validation_type,
                expected_result=testcase.expected_result,
                response=poll_response,
                validation_rules=self._parse_json_object(testcase.validation_rules),
            )

            return {
                "tc_id": testcase.tc_id,
                "service": testcase.service,
                "description": testcase.description,
                "bank": testcase.bank,
                "api_name": f"{initial_api_name}->{polling_definition.get('api', 'polling')}",
                "test_type": testcase.test_type,
                "status": "PASSED" if validation["passed"] else "FAILED",
                "expected": testcase.expected_result,
                "actual": validation["actual"],
                "status_code": poll_response.get("status_code"),
                "response_time_ms": poll_response.get("elapsed_ms"),
                "request_payload": {
                    "initial": initial_payload,
                    "polling": poll_response.get("request_payload", {}),
                },
                "request_headers": {
                    "initial": initial_headers,
                    "polling": poll_response.get("request_headers", {}),
                },
                "response_body": {
                    "initial": initial_response.get("body"),
                    "final": poll_response.get("body"),
                },
                "error": "",
            }

        # Handle conditional chained flow with on_success
        on_success = flow.get("on_success") or flow.get("next")
        if on_success:
            # Determine success by checking response_code field in initial response body
            initial_body = initial_response.get("body", {}) or {}
            # Try common locations for response code (top-level or nested under 'response')
            resp_code = self._get_path_value(initial_body, "response_code")
            if resp_code is None:
                resp_code = self._get_path_value(initial_body, "response.response_code")

            if str(resp_code) == "00":
                # extract mappings into tokens for next request
                extra_tokens: Dict[str, str] = {}
                mappings = on_success.get("mappings", [])
                for m in mappings:
                    from_path = m.get("from")
                    to_token = m.get("to")
                    if not from_path or not to_token:
                        continue

                    # Try exact path, then try with 'response.' prefix for nested responses
                    val = self._get_path_value(initial_body, from_path)
                    if val is None:
                        val = self._get_path_value(initial_body, f"response.{from_path}")

                    if val is not None:
                        extra_tokens[str(to_token).upper()] = str(val)

                next_api_name = on_success.get("next_api")
                next_request_file = on_success.get("next_request_file")
                if not next_api_name or not next_request_file:
                    raise ValueError("on_success flow must define next_api and next_request_file")

                next_api_definition = self._get_api_definition(next_api_name, testcase.service)
                next_payload = self._render_payload(next_request_file, extra_tokens=extra_tokens, service=testcase.service)
                next_url = self._build_url(bank_base_url, next_api_definition["endpoint"])
                next_headers = self._build_request_headers(next_api_definition, next_api_name, testcase.service)
                next_response = self.api_client.call(
                    method=str(next_api_definition["method"]).upper(),
                    url=next_url,
                    payload=next_payload,
                    headers=next_headers,
                )

                validation = self.validator.validate(
                    validation_type=testcase.validation_type,
                    expected_result=testcase.expected_result,
                    response=next_response,
                    validation_rules=self._parse_json_object(testcase.validation_rules),
                )

                return {
                    "tc_id": testcase.tc_id,
                    "service": testcase.service,
                    "description": testcase.description,
                    "bank": testcase.bank,
                    "api_name": f"{initial_api_name}->{next_api_name}",
                    "test_type": testcase.test_type,
                    "status": "PASSED" if validation["passed"] else "FAILED",
                    "expected": testcase.expected_result,
                    "actual": validation["actual"],
                    "status_code": next_response.get("status_code"),
                    "response_time_ms": next_response.get("elapsed_ms"),
                    "request_payload": {
                        "initial": initial_payload,
                        "next": next_payload,
                    },
                    "request_headers": {
                        "initial": initial_headers,
                        "next": next_headers,
                    },
                    "response_body": {
                        "initial": initial_response.get("body"),
                        "final": next_response.get("body"),
                    },
                    "error": "",
                }
            else:
                # negative branch: initial lookup failed; do not call payment
                validation = self.validator.validate(
                    validation_type=testcase.validation_type,
                    expected_result=testcase.expected_result,
                    response=initial_response,
                    validation_rules=self._parse_json_object(testcase.validation_rules),
                )

                return {
                    "tc_id": testcase.tc_id,
                    "service": testcase.service,
                    "description": testcase.description,
                    "bank": testcase.bank,
                    "api_name": initial_api_name,
                    "test_type": testcase.test_type,
                    "status": "PASSED" if validation["passed"] else "FAILED",
                    "expected": testcase.expected_result,
                    "actual": validation["actual"],
                    "status_code": initial_response.get("status_code"),
                    "response_time_ms": initial_response.get("elapsed_ms"),
                    "request_payload": {
                        "initial": initial_payload,
                    },
                    "request_headers": {
                        "initial": initial_headers,
                    },
                    "response_body": {
                        "initial": initial_response.get("body"),
                    },
                    "error": "",
                }

        # If flow did not match any known pattern
        raise ValueError(f"Unsupported flow configuration for '{testcase.flow_config}' in service '{testcase.service}'")

    def _poll_for_final_status(
        self,
        bank_base_url: str,
        polling_definition: Dict[str, Any],
        initial_response: Dict[str, Any],
        service: str = "cas",
    ) -> Dict[str, Any]:
        """Poll for final status with service context awareness."""
        polling_api_name = polling_definition.get("api")
        if not polling_api_name:
            raise ValueError("Polling API name is required in flow configuration")

        api_definition = self._get_api_definition(polling_api_name, service)
        interval_seconds = int(polling_definition.get("interval_seconds", 5))
        timeout_seconds = int(polling_definition.get("timeout_seconds", 60))
        deadline = time.time() + timeout_seconds

        request_file = polling_definition.get("request_file", "")
        correlation_source_path = polling_definition.get("correlation_from_response_json_path", "")
        correlation_target_path = polling_definition.get("correlation_in_request_json_path", "")
        correlation_value = self._get_path_value(initial_response.get("body", {}), correlation_source_path)

        while time.time() <= deadline:
            extra_tokens = {}
            if correlation_value:
                extra_tokens["TRANSACTION_ID"] = str(correlation_value)

            polling_payload = (
                self._render_payload(request_file, extra_tokens=extra_tokens, service=service)
                if request_file
                else {}
            )

            if correlation_target_path and correlation_value is not None:
                self._set_path_value(polling_payload, correlation_target_path, correlation_value)

            poll_url = self._build_url(bank_base_url, api_definition["endpoint"])
            polling_headers = self._build_request_headers(api_definition, polling_api_name, service)
            response = self.api_client.call(
                method=str(api_definition["method"]).upper(),
                url=poll_url,
                payload=polling_payload,
                headers=polling_headers,
            )
            response["request_payload"] = polling_payload
            response["request_headers"] = polling_headers

            if self._is_flow_stop_condition_met(response, polling_definition.get("stop_when", {})):
                return response

            time.sleep(interval_seconds)

        response = {
            "status_code": None,
            "body": {"error": "Polling timeout reached without terminal status"},
            "text": "Polling timeout reached without terminal status",
            "elapsed_ms": timeout_seconds * 1000,
            "headers": {},
            "request_payload": {},
        }
        return response

    def _is_flow_stop_condition_met(self, response: Dict[str, Any], stop_when: Dict[str, Any]) -> bool:
        condition_type = stop_when.get("type", "response")

        if condition_type == "json_field":
            field_path = stop_when.get("field", "")
            allowed_values = stop_when.get("allowed_values", [])
            actual_value = self._get_path_value(response.get("body", {}), field_path)
            return str(actual_value) in [str(value) for value in allowed_values]

        if condition_type == "status_code":
            allowed_codes = stop_when.get("allowed_values", [])
            return response.get("status_code") in allowed_codes

        contains = str(stop_when.get("contains", "")).strip()
        return bool(contains and contains in str(response.get("text", "")))

    def _render_payload(self, request_file_name: str, extra_tokens: Dict[str, str], service: str = "cas") -> Dict[str, Any]:
        if not request_file_name:
            return {}
        request_file_path = self.requests_dir / service / request_file_name
        if not request_file_path.exists():
            request_file_path = self.requests_dir / request_file_name
        template = TemplateEngine.load_json_template(request_file_path)
        values = DataGenerator.default_tokens()
        values.update(self.test_data_tokens)   # user-defined variables from test_data.yaml
        values.update(extra_tokens)            # call-site overrides take highest priority
        rendered = TemplateEngine.render_template(template, values)
        return rendered

    def _build_request_headers(
        self, api_definition: Dict[str, Any], api_name: str = "", service: str = "cas"
    ) -> Dict[str, str]:
        """Build request headers with service-aware token injection."""
        headers = dict(api_definition.get("headers", {}))
        token_value = self._get_token_for_api(api_name, service) if api_name else ""
        if token_value:
            headers["Authorization"] = f"Bearer {token_value}"
        return headers

    def _capture_auth_token_if_configured(self, testcase: ServiceTestCase, response: Dict[str, Any]) -> None:
        """Capture and store auth token with service context awareness."""
        auth_api_names = self._auth_api_names(testcase.service)
        if testcase.api_name not in auth_api_names:
            return

        auth_config = self.service_manager.get_service_auth_config(testcase.service)
        if not auth_config:
            # Backward compatibility for legacy global auth config.
            auth_config = self.api_config.get("auth", {})
        if not auth_config:
            return

        token_path = str(auth_config.get("token_json_path", "token")).strip()
        token_value = self._get_path_value(response.get("body", {}), token_path)
        if token_value is None:
            self.framework_logger.warning(
                "[%s] Auth API executed but token not found at path '%s' for tc_id=%s",
                testcase.service,
                token_path,
                testcase.tc_id,
            )
            return

        token_str = str(token_value)
        self.runtime_context["auth_token"] = token_str

        channel_key = self._extract_channel_key(testcase)
        if channel_key:
            self.service_token_context[testcase.service][channel_key] = token_str
            self.framework_logger.info(
                "[%s] Auth token captured from tc_id=%s for channel=%s",
                testcase.service,
                testcase.tc_id,
                channel_key,
            )
        else:
            self.service_token_context[testcase.service]["default"] = token_str
            self.framework_logger.info(
                "[%s] Auth token captured from tc_id=%s as default token", testcase.service, testcase.tc_id
            )

    def _extract_channel_key(self, testcase: ServiceTestCase) -> str:
        """Extract channel key from request file payload."""
        try:
            # Prefer service-scoped payload over root-level payload when both exist.
            payload_path_options = [
                self.requests_dir / testcase.service / testcase.request_file,
                self.requests_dir / testcase.request_file,
            ]
            
            for payload_path in payload_path_options:
                if payload_path.exists():
                    with payload_path.open("r", encoding="utf-8") as handle:
                        payload = json.load(handle)
                        channel = str(payload.get("channel", "")).strip()
                        if channel:
                            return channel
        except (json.JSONDecodeError, FileNotFoundError):
            pass
        return ""

    def _auth_api_names(self, service: str = "cas") -> List[str]:
        """Get authentication API names for a service."""
        auth_config = self.service_manager.get_service_auth_config(service)
        if not auth_config:
            # Fallback to global auth config for backward compatibility
            auth_config = self.api_config.get("auth", {})
        
        configured_names = auth_config.get("apis", ["authenticate"])
        if not isinstance(configured_names, list):
            return ["authenticate"]
        return [str(name).strip() for name in configured_names if str(name).strip()]

    def _resolve_service_for_row(
        self,
        row: Any,
        header_index: Dict[str, int],
        default_service: str = "cas",
    ) -> str:
        """Resolve service for a testcase row.

        Priority:
        1. Explicit Service column when present.
        2. Request file name hints.
        3. API name unique to one configured service.
        4. Source default fallback.
        """
        if "Service" in header_index:
            service_value = str(row[header_index["Service"]] or "").strip().lower()
            if service_value:
                return service_value

        request_file = str(row[header_index.get("Request_File", -1)] or "").strip().lower()
        api_name = str(row[header_index.get("API_Name", -1)] or "").strip().lower()

        request_hints = [
            ("acq", "acquiring"),
            ("acquiring", "acquiring"),
            ("mcas", "merchant_cas"),
            ("merchant", "merchant_cas"),
            ("merchant_cas", "merchant_cas"),
        ]
        for token, service_id in request_hints:
            if token in request_file:
                return service_id

        matching_services = []
        for service_id in self.service_manager.get_services_by_status(enabled_only=False):
            service_apis = self.service_manager.get_service_apis(service_id)
            if api_name and api_name in service_apis:
                matching_services.append(service_id)

        if len(matching_services) == 1:
            return matching_services[0]

        return default_service

    def _get_bank_base_url(self, bank_name: str, service: str = "cas") -> str:
        """Get base URL with service priority, then bank fallback mapping."""
        service_base_url = self.service_manager.get_service_base_url(service)
        if service_base_url:
            return service_base_url

        banks = self.server_config.get("banks", {})
        if bank_name not in banks:
            raise ValueError(f"Unknown bank in servers.yaml: {bank_name}")
        
        bank_config = banks[bank_name]
        
        # If bank mapping includes service, use that
        if isinstance(bank_config, dict) and "service" in bank_config:
            service = bank_config.get("service", "cas")
            base_url = self.service_manager.get_service_base_url(service)
            if base_url:
                return base_url
        
        # Otherwise use direct base_url from bank config
        base_url = str(bank_config.get("base_url", "") if isinstance(bank_config, dict) else bank_config).strip()
        if not base_url:
            raise ValueError(f"Empty base_url for bank: {bank_name}")
        return base_url

    def _get_api_definition(self, api_name: str, service: str = "cas") -> Dict[str, Any]:
        """Get API definition with service context awareness."""
        # Try service-scoped API first
        service_apis = self.service_manager.get_service_apis(service)
        if api_name in service_apis:
            api_def = service_apis[api_name]
            if "endpoint" in api_def and "method" in api_def:
                return api_def
        
        # Fall back to global APIs for backward compatibility
        global_apis = self.api_config.get("apis", {})
        if api_name in global_apis:
            api_def = global_apis[api_name]
            if "endpoint" in api_def and "method" in api_def:
                return api_def
        
        raise ValueError(f"Unknown API '{api_name}' in service '{service}' or global config")

    @staticmethod
    def _build_url(base_url: str, endpoint: str) -> str:
        return base_url.rstrip("/") + "/" + endpoint.lstrip("/")

    @staticmethod
    def _tc_order_key(tc_id: str) -> int:
        match = re.search(r"(\d+)", str(tc_id))
        if not match:
            return 10**9
        return int(match.group(1))

    @classmethod
    def _is_sequential_testcase(cls, tc_id: str) -> bool:
        return cls._tc_order_key(tc_id) >= 11

    @staticmethod
    def _parse_json_object(value: str) -> Dict[str, Any]:
        if not value:
            return {}
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}

    @staticmethod
    def _get_path_value(payload: Any, path: str) -> Any:
        if not path:
            return None
        current = payload
        for part in path.split("."):
            if isinstance(current, dict) and part in current:
                current = current[part]
            else:
                return None
        return current

    @staticmethod
    def _set_path_value(payload: Dict[str, Any], path: str, value: Any) -> None:
        if not path:
            return
        parts = path.split(".")
        current = payload
        for part in parts[:-1]:
            if part not in current or not isinstance(current[part], dict):
                current[part] = {}
            current = current[part]
        current[parts[-1]] = value

    def _write_results_json(self, run_tag: str, results: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
        output = {"run_tag": run_tag, "summary": summary, "results": results}
        output_file = self.results_dir / f"results_{run_tag}.json"
        with output_file.open("w", encoding="utf-8") as handle:
            json.dump(output, handle, ensure_ascii=False, indent=2)

    def _write_results_csv(self, run_tag: str, results: List[Dict[str, Any]]) -> None:
        output_file = self.results_dir / f"results_{run_tag}.csv"
        headers = [
            "TC_ID",
            "Service",
            "Description",
            "Test_Type",
            "Status",
            "Expected",
            "Actual",
            "Response_Time_ms",
            "Status_Code",
            "API_Name",
            "Bank",
        ]
        lines = [",".join(headers)]

        for item in results:
            row = [
                self._safe_csv(item.get("tc_id")),
                self._safe_csv(item.get("service", "cas")),
                self._safe_csv(item.get("description")),
                self._safe_csv(item.get("test_type", "positive")),
                self._safe_csv(item.get("status")),
                self._safe_csv(item.get("expected")),
                self._safe_csv(item.get("actual")),
                self._safe_csv(item.get("response_time_ms")),
                self._safe_csv(item.get("status_code")),
                self._safe_csv(item.get("api_name")),
                self._safe_csv(item.get("bank")),
            ]
            lines.append(",".join(row))

        output_file.write_text("\n".join(lines), encoding="utf-8")

    def _extract_failure_reason(self, item: Dict[str, Any]) -> str:
        if str(item.get("status", "")).upper() == "PASSED":
            return "N/A"

        def first_nested_value(obj: Any, keys: List[str]) -> str:
            if isinstance(obj, dict):
                for key in keys:
                    value = obj.get(key)
                    if value not in (None, ""):
                        return str(value)
                for value in obj.values():
                    found = first_nested_value(value, keys)
                    if found:
                        return found
            elif isinstance(obj, list):
                for value in obj:
                    found = first_nested_value(value, keys)
                    if found:
                        return found
            return ""

        error_text = str(item.get("error") or "").strip()
        if error_text:
            return error_text

        response_reason = first_nested_value(
            item.get("response_body"),
            ["response_desc", "message", "error", "detail", "response_code"],
        )
        if response_reason:
            return response_reason

        actual = item.get("actual", "")
        if isinstance(actual, str):
            stripped = actual.strip()
            if stripped:
                try:
                    parsed_actual = json.loads(stripped)
                    actual_reason = first_nested_value(
                        parsed_actual,
                        ["response_desc", "message", "error", "detail", "response_code"],
                    )
                    if actual_reason:
                        return actual_reason
                except json.JSONDecodeError:
                    return stripped[:160]
        elif actual not in (None, ""):
            return str(actual)[:160]

        return "No failure reason captured"

    def _build_failure_analysis(self, results: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        failures = [item for item in results if str(item.get("status", "")).upper() != "PASSED"]

        by_api: Dict[str, int] = {}
        by_reason: Dict[str, int] = {}
        for item in failures:
            api_name = str(item.get("api_name") or "Unknown")
            reason = self._extract_failure_reason(item)
            by_api[api_name] = by_api.get(api_name, 0) + 1
            by_reason[reason] = by_reason.get(reason, 0) + 1

        slowest = sorted(
            failures,
            key=lambda item: int(item.get("response_time_ms") or 0),
            reverse=True,
        )[:8]

        return {
            "by_api": [
                {"label": label, "count": count}
                for label, count in sorted(by_api.items(), key=lambda row: (-row[1], row[0]))
            ],
            "by_reason": [
                {"label": label, "count": count}
                for label, count in sorted(by_reason.items(), key=lambda row: (-row[1], row[0]))
            ],
            "slowest": [
                {
                    "tc_id": str(item.get("tc_id", "")),
                    "api_name": str(item.get("api_name", "")),
                    "reason": self._extract_failure_reason(item),
                    "response_time_ms": int(item.get("response_time_ms") or 0),
                }
                for item in slowest
            ],
        }

    def _redact_for_report(self, value: Any, key_hint: str = "") -> Any:
        sensitive_keys = {
            "authorization",
            "password",
            "token",
            "access_token",
            "refresh_token",
            "secret",
            "api_key",
        }
        key_lower = key_hint.lower()

        if isinstance(value, dict):
            redacted: Dict[str, Any] = {}
            for key, item in value.items():
                if str(key).lower() in sensitive_keys:
                    redacted[key] = "[REDACTED]"
                else:
                    redacted[key] = self._redact_for_report(item, str(key))
            return redacted

        if isinstance(value, list):
            return [self._redact_for_report(item, key_hint) for item in value]

        if isinstance(value, str):
            if key_lower in sensitive_keys:
                return "[REDACTED]"
            return self._redact_text_for_report(value)

        return value

    @staticmethod
    def _redact_text_for_report(text: str) -> str:
        if not text:
            return text

        redacted = re.sub(r"Bearer\s+[A-Za-z0-9._~+/=-]+", "Bearer [REDACTED]", text, flags=re.IGNORECASE)
        redacted = re.sub(
            r"\b[A-Za-z0-9_-]{20,}\.[A-Za-z0-9_-]{20,}\.[A-Za-z0-9_-]{20,}\b",
            "[REDACTED_JWT]",
            redacted,
        )
        redacted = re.sub(
            r"\bPK\d{2}[A-Z0-9]{4,30}\b",
            lambda match: f"{match.group(0)[:6]}...[REDACTED]...{match.group(0)[-4:]}",
            redacted,
        )
        redacted = re.sub(
            r"\b\d{10,}\b",
            lambda match: f"{match.group(0)[:3]}...[REDACTED]...{match.group(0)[-3:]}",
            redacted,
        )
        return redacted

    def _write_html_report(
        self,
        run_tag: str,
        results: List[Dict[str, Any]],
        summary: Dict[str, Any],
        started_at: datetime,
        finished_at: datetime,
    ) -> None:
        report_file = self.reports_dir / f"report_{run_tag}.html"

        pass_rate = float(summary.get("pass_percent", 0.0))
        hide_sensitive_data = self._reporting_bool("html_report.hide_sensitive_data", True)
        show_failure_analysis = self._reporting_bool("html_report.show_failure_analysis", True)
        show_details = self._reporting_bool("html_report.show_details", True)
        initial_analysis_hidden = "true" if not show_failure_analysis else "false"
        initial_details_hidden = "true" if not show_details else "false"
        analysis_section_class = "analysis-section" if show_failure_analysis else "analysis-section analysis-hidden"
        details_wrap_class = "details-wrap" if show_details else "details-wrap details-hidden"
        duration_seconds = max(0.0, (finished_at - started_at).total_seconds())
        average_response_ms = (
            round(
                sum(int(item.get("response_time_ms") or 0) for item in results)
                / len(results),
                2,
            )
            if results
            else 0
        )
        failure_analysis = self._build_failure_analysis(results)

        def report_value(value: Any) -> Any:
            return self._redact_for_report(value) if hide_sensitive_data else value

        def report_text(value: Any) -> str:
            text = str(value if value is not None else "")
            return self._redact_text_for_report(text) if hide_sensitive_data else text

        def render_analysis_list(
            rows: List[Dict[str, Any]],
            label_key: str = "label",
            action_type: str = "",
        ) -> str:
            if not rows:
                return "<li class='analysis-empty'>No failures captured</li>"
            return "".join(
                "<li>"
                f"<button type='button' class='analysis-action' "
                f"data-analysis-action='{self._html_escape(action_type)}' "
                f"data-analysis-value='{self._html_escape(report_text(row.get(label_key, '')))}'>"
                f"<span>{self._html_escape(report_text(row.get(label_key, '')))}</span>"
                "</button>"
                f"<strong>{self._html_escape(str(row.get('count', '')))}</strong>"
                "</li>"
                for row in rows
            )

        def render_slowest_list(rows: List[Dict[str, Any]]) -> str:
            if not rows:
                return "<li class='analysis-empty'>No slow failed cases</li>"
            return "".join(
                "<li>"
                f"<button type='button' class='analysis-action' "
                "data-analysis-action='tc' "
                f"data-analysis-value='{self._html_escape(str(row.get('tc_id', '')))}'>"
                f"<span><code>{self._html_escape(str(row.get('tc_id', '')))}</code> "
                f"{self._html_escape(str(row.get('api_name', '')))} - "
                f"{self._html_escape(report_text(row.get('reason', '')))}</span>"
                "</button>"
                f"<strong>{self._html_escape(str(row.get('response_time_ms', 0)))} ms</strong>"
                "</li>"
                for row in rows
            )

        table_rows = []
        details_sections = []
        service_nav: Dict[str, Dict[str, Any]] = {}
        api_names = set()
        test_types = set()
        for index, item in enumerate(results, start=1):
            status = str(item.get("status", "")).upper()
            status_class = "status-pass" if status == "PASSED" else "status-fail"
            test_type_display = str(item.get("test_type", "positive")).capitalize()
            test_type_class = "test-positive" if test_type_display.lower() == "positive" else "test-negative"
            details_id = f"details-{run_tag}-{index}"
            actual_modal_id = f"actual-modal-{run_tag}-{index}"
            service_id = str(item.get("service", "cas")).strip() or "cas"
            service_key = self._html_escape(service_id)
            api_name = str(item.get("api_name", ""))
            api_names.add(api_name)
            test_types.add(str(item.get("test_type", "positive")).lower())
            response_ms = int(item.get("response_time_ms") or 0)
            response_class = "rt-fast" if response_ms <= 1000 else "rt-medium" if response_ms <= 3000 else "rt-slow"
            failure_reason = self._extract_failure_reason(item)
            failure_reason_display = report_text(failure_reason)
            failure_reason_class = "reason-muted" if status == "PASSED" else "reason-fail"
            search_blob = " ".join(
                [
                    str(item.get("tc_id", "")),
                    service_id,
                    str(item.get("description", "")),
                    str(item.get("bank", "")),
                    api_name,
                    status,
                    str(item.get("expected", "")),
                    failure_reason_display,
                ]
            ).lower()

            if service_id not in service_nav:
                display_name = service_id
                service_config = self.service_manager.get_service(service_id)
                if service_config:
                    display_name = service_config.display_name
                service_nav[service_id] = {
                    "display_name": display_name,
                    "passed": 0,
                    "failed": 0,
                    "cases": [],
                }
            if status == "PASSED":
                service_nav[service_id]["passed"] += 1
            else:
                service_nav[service_id]["failed"] += 1
            service_nav[service_id]["cases"].append(
                {
                    "tc_id": str(item.get("tc_id", "")),
                    "description": str(item.get("description", "")),
                    "details_id": details_id,
                }
            )

            actual_value = item.get("actual", "")
            actual_text = str(report_value(actual_value))
            actual_display = self._format_actual_for_modal(actual_text)

            table_rows.append(
                f"<tr class='result-row' data-service='{service_key}' "
                f"data-row-status='{status.lower()}' "
                f"data-api='{self._html_escape(api_name)}' "
                f"data-test-type='{self._html_escape(str(item.get('test_type', 'positive')).lower())}' "
                f"data-search='{self._html_escape(search_blob)}'>"
                f"<td><code>{self._html_escape(str(item.get('tc_id', '')))}</code></td>"
                f"<td>{service_key}</td>"
                f"<td>{self._html_escape(str(item.get('description', '')))}</td>"
                f"<td><span class='type-chip {test_type_class}'>{self._html_escape(test_type_display)}</span></td>"
                f"<td>{self._html_escape(str(item.get('bank', '')))}</td>"
                f"<td>{self._html_escape(api_name)}</td>"
                f"<td><span class='status-chip {status_class}'>{self._html_escape(status)}</span></td>"
                f"<td>{self._html_escape(str(item.get('expected', '')))}</td>"
                f"<td><button type='button' class='actual-btn' data-modal-target='{actual_modal_id}'>View</button></td>"
                f"<td><span class='rt-chip {response_class}'>{response_ms}</span></td>"
                f"<td><a class='detail-link' href='#{details_id}'>View</a></td>"
                f"<td><span class='reason-chip {failure_reason_class}'>{self._html_escape(failure_reason_display)}</span></td>"
                "</tr>"
            )

            details_sections.append(
                f"<div id='{actual_modal_id}' class='modal-overlay' aria-hidden='true'>"
                "<div class='modal-card' role='dialog' aria-modal='true' aria-labelledby='actual-modal-title'>"
                "<div class='modal-head'>"
                f"<h3 id='actual-modal-title'>{self._html_escape(str(item.get('tc_id', '')))} Actual Response</h3>"
                "<button type='button' class='modal-close' data-modal-close='true' aria-label='Close'>&times;</button>"
                "</div>"
                "<div class='modal-meta'>"
                f"<span>API: {self._html_escape(str(item.get('api_name', '')))}</span>"
                f"<span>Status: {self._html_escape(status)}</span>"
                "</div>"
                f"<pre class='modal-pre'>{self._html_escape(actual_display)}</pre>"
                "</div>"
                "</div>"
            )

            # Render chained flow steps explicitly when present
            req_payload = item.get('request_payload', {}) or {}
            resp_body = item.get('response_body', {}) or {}

            def render_json(obj):
                safe_obj = report_value(obj)
                return self._html_escape(json.dumps(safe_obj, indent=2, ensure_ascii=False))

            sections_html = []
            # Common: Request Headers
            sections_html.append(
                "<article>"
                "<h4>Request Headers</h4>"
                f"<pre>{render_json(item.get('request_headers', {}))}</pre>"
                "</article>"
            )

            # If chained flow with initial/next
            if isinstance(req_payload, dict) and ('initial' in req_payload or 'next' in req_payload or 'polling' in req_payload):
                # Step 1
                sections_html.append(
                    "<article>"
                    f"<h4>Step 1 Request ({self._html_escape(str(item.get('api_name', '').split('->')[0]))})</h4>"
                    f"<pre>{render_json(req_payload.get('initial', req_payload))}</pre>"
                    "</article>"
                )

                # Step 1 Response
                first_resp = resp_body.get('initial', resp_body.get('initial', {}))
                sections_html.append(
                    "<article>"
                    f"<h4>Step 1 Response ({self._html_escape(str(item.get('api_name', '').split('->')[0]))})</h4>"
                    f"<pre>{render_json(first_resp)}</pre>"
                    "</article>"
                )

                # Step 2 (next)
                if 'next' in req_payload:
                    sections_html.append(
                        "<article>"
                        f"<h4>Step 2 Request ({self._html_escape(str(item.get('api_name', '').split('->')[-1]))})</h4>"
                        f"<pre>{render_json(req_payload.get('next', {}))}</pre>"
                        "</article>"
                    )
                    sections_html.append(
                        "<article class='span-2'>"
                        f"<h4>Step 2 Response ({self._html_escape(str(item.get('api_name', '').split('->')[-1]))})</h4>"
                        f"<pre>{render_json(resp_body.get('final', {}))}</pre>"
                        "</article>"
                    )
                elif 'polling' in req_payload:
                    sections_html.append(
                        "<article class='span-2'>"
                        "<h4>Polling Final Response</h4>"
                        f"<pre>{render_json(resp_body.get('final', {}))}</pre>"
                        "</article>"
                    )
                else:
                    # fallback: show whole payload/response
                    sections_html.append(
                        "<article class='span-2'>"
                        "<h4>Full Request Payload</h4>"
                        f"<pre>{render_json(req_payload)}</pre>"
                        "</article>"
                    )
                    sections_html.append(
                        "<article class='span-2'>"
                        "<h4>Full Response Body</h4>"
                        f"<pre>{render_json(resp_body)}</pre>"
                        "</article>"
                    )
            else:
                # Non-chained: show request payload and response body
                sections_html.append(
                    "<article>"
                    "<h4>Request Payload</h4>"
                    f"<pre>{render_json(req_payload)}</pre>"
                    "</article>"
                )
                sections_html.append(
                    "<article class='span-2'>"
                    "<h4>Response Body</h4>"
                    f"<pre>{render_json(resp_body)}</pre>"
                    "</article>"
                )

            details_sections.append(
                f"<section id='{details_id}' class='detail-card result-detail' "
                f"data-service='{service_key}' "
                f"data-detail-status='{status.lower()}' "
                f"data-api='{self._html_escape(api_name)}' "
                f"data-test-type='{self._html_escape(str(item.get('test_type', 'positive')).lower())}' "
                f"data-search='{self._html_escape(search_blob)}'>"
                "<div class='detail-head'>"
                f"<h3>{self._html_escape(str(item.get('tc_id', '')))} - {self._html_escape(str(item.get('description', '')))}</h3>"
                f"<div class='detail-meta'>Service: {service_key} | Test Type: {self._html_escape(test_type_display)} | API: {self._html_escape(api_name)} | Bank: {self._html_escape(str(item.get('bank', '')))} | Status: <span class='{status_class}'>{self._html_escape(status)}</span> | Failure Reason: {self._html_escape(failure_reason_display)}</div>"
                "</div>"
                "<div class='json-grid'>"
                f"{''.join(sections_html)}"
                "</div>"
                "</section>"
            )

        service_items = []
        total_passed = sum(info["passed"] for info in service_nav.values())
        total_cases = sum(info["passed"] + info["failed"] for info in service_nav.values())
        service_items.append(
            "<li class='service-item'>"
            "<button class='service-link active' type='button' data-service-filter=''>"
            "<span class='service-name'>All Services</span>"
            f"<span class='service-stats'><span class='svc-pass'>{total_passed}</span>/<span class='svc-total'>{total_cases}</span></span>"
            "</button>"
            "</li>"
        )
        for service_id, info in service_nav.items():
            cases_html = []
            for case in info["cases"]:
                tc_id = self._html_escape(case["tc_id"])
                description = self._html_escape(case["description"])
                details_id = self._html_escape(case["details_id"])
                cases_html.append(
                    f"<li><a href='#{details_id}'>{tc_id} - {description}</a></li>"
                )

            service_items.append(
                "<li class='service-item'>"
                f"<button class='service-link' type='button' data-service-filter='{self._html_escape(service_id)}'>"
                f"<span class='service-name'>{self._html_escape(info['display_name'])}</span>"
                f"<span class='service-stats'><span class='svc-pass'>{info['passed']}</span>/<span class='svc-total'>{info['passed'] + info['failed']}</span></span>"
                "</button>"
                f"<ul class='service-cases'>{''.join(cases_html)}</ul>"
                "</li>"
            )

        sidebar_html = (
            "<aside class='service-sidebar'>"
            "<h2>Services</h2>"
            "<ul class='service-list'>"
            f"{''.join(service_items)}"
            "</ul>"
            "</aside>"
        )
        api_options = "".join(
            f"<option value='{self._html_escape(api)}'>{self._html_escape(api or 'Unknown')}</option>"
            for api in sorted(api_names)
        )
        test_type_options = "".join(
            f"<option value='{self._html_escape(test_type)}'>{self._html_escape(test_type.capitalize())}</option>"
            for test_type in sorted(test_types)
        )
        failure_analysis_html = f"""
        <section class="analysis-grid" aria-label="Failure Analysis">
            <article class="analysis-card">
                <h3>Failures by API</h3>
                <ul>{render_analysis_list(failure_analysis["by_api"], action_type="api")}</ul>
            </article>
            <article class="analysis-card">
                <h3>Top Failure Reasons</h3>
                <ul>{render_analysis_list(failure_analysis["by_reason"], action_type="reason")}</ul>
            </article>
            <article class="analysis-card">
                <h3>Slowest Failed Cases</h3>
                <ul>{render_slowest_list(failure_analysis["slowest"])}</ul>
            </article>
        </section>
        """

        html = f"""
<!doctype html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>SIT Automation Report - {run_tag}</title>
  <style>
        :root {{
            --bg-1: #f3f6fb;
            --bg-2: #e9f0fb;
            --surface: #ffffff;
            --surface-2: #f9fbff;
            --ink: #0f172a;
            --muted: #475569;
            --line: #d9e2ef;
            --brand: #0f4c81;
            --brand-soft: #dbeafe;
            --pass-bg: #dcfce7;
            --pass-fg: #14532d;
            --fail-bg: #fee2e2;
            --fail-fg: #7f1d1d;
            --fast-bg: #dbeafe;
            --fast-fg: #1e3a8a;
            --mid-bg: #fef3c7;
            --mid-fg: #92400e;
            --slow-bg: #fde68a;
            --slow-fg: #78350f;
            --shadow: 0 8px 28px rgba(15, 23, 42, 0.08);
        }}
        * {{ box-sizing: border-box; }}
        body {{
            margin: 0;
            color: var(--ink);
            font-family: "Segoe UI", "Noto Sans", Tahoma, sans-serif;
            background:
                radial-gradient(1200px 400px at 90% -10%, #cfe3ff 0%, transparent 50%),
                linear-gradient(180deg, var(--bg-2) 0%, var(--bg-1) 52%, #f8fbff 100%);
            min-height: 100vh;
            line-height: 1.45;
        }}
        .container {{ max-width: 1480px; margin: 0 auto; padding: 24px; }}
        .layout {{ display: grid; grid-template-columns: 320px 1fr; gap: 16px; align-items: start; }}
	.layout.sidebar-collapsed {{ grid-template-columns: 1fr; }}
        .service-sidebar {{
            position: sticky;
            top: 14px;
            background: var(--surface);
            border: 1px solid var(--line);
            border-radius: 12px;
            box-shadow: var(--shadow);
            padding: 12px;
            max-height: calc(100vh - 28px);
            overflow: auto;
        }}
        .layout.sidebar-collapsed .service-sidebar {{ display: none; }}
        .service-sidebar h2 {{ margin: 2px 4px 10px; font-size: 16px; color: #0b3f66; }}
        .service-list {{ list-style: none; margin: 0; padding: 0; display: grid; gap: 10px; }}
        .service-item {{ border: 1px solid #dde7f5; border-radius: 10px; background: #f8fbff; overflow: hidden; }}
        .service-link {{
            width: 100%;
            border: none;
            background: transparent;
            padding: 10px;
            cursor: pointer;
            display: flex;
            justify-content: space-between;
            align-items: center;
            text-align: left;
            font-weight: 700;
            color: #0b3f66;
        }}
        .service-link.active {{ background: #e7f0fb; }}
        .service-name {{ font-size: 13px; }}
        .service-stats {{ font-size: 12px; color: #334155; }}
        .svc-pass {{ color: #14532d; font-weight: 800; }}
        .svc-total {{ color: #0f172a; }}
        .service-cases {{ list-style: none; margin: 0; padding: 0 10px 10px; display: grid; gap: 4px; }}
        .service-cases a {{ color: #0a4d86; text-decoration: none; font-size: 12px; }}
        .service-cases a:hover {{ text-decoration: underline; }}
        .content {{ min-width: 0; }}
        .toolbar {{ display: flex; align-items: center; justify-content: space-between; margin-bottom: 10px; flex-wrap: wrap; gap: 8px; }}
        .filter-panel {{ display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }}
        .toolbar-actions {{ display: flex; flex-wrap: wrap; gap: 8px; align-items: center; }}
        .status-filters {{ display: flex; gap: 6px; }}
        .filter-btn {{
            border: 1px solid #c7d7ec;
            background: #f0f5ff;
            color: #334155;
            border-radius: 8px;
            padding: 6px 14px;
            font-weight: 600;
            cursor: pointer;
            font-size: 12px;
        }}
        .filter-btn:hover {{ background: #dbeafe; border-color: #93c5fd; }}
        .filter-btn.active {{ background: #0f4c81; color: #fff; border-color: #0f4c81; }}
        .filter-btn.pass-btn.active {{ background: #14532d; border-color: #14532d; }}
        .filter-btn.fail-btn.active {{ background: #7f1d1d; border-color: #7f1d1d; }}
        .filter-select,
        .search-input {{
            border: 1px solid #c7d7ec;
            background: #ffffff;
            color: #1e293b;
            border-radius: 8px;
            padding: 7px 10px;
            font-size: 12px;
            min-height: 32px;
        }}
        .filter-select {{ min-width: 150px; }}
        .search-input {{ min-width: 230px; }}
        .utility-btn,
        .sidebar-toggle {{
            border: 1px solid #a7c2e8;
            background: #eef4ff;
            color: #0a4d86;
            border-radius: 8px;
            padding: 7px 12px;
            font-weight: 700;
            cursor: pointer;
            font-size: 12px;
        }}
        .utility-btn:hover,
        .sidebar-toggle:hover {{
            background: #dbeafe;
            border-color: #8ab0e0;
        }}
        .header {{
            background: linear-gradient(135deg, #0f4c81 0%, #0a6aa6 100%);
            color: #ffffff;
            border-radius: 14px;
            padding: 22px 24px;
            box-shadow: var(--shadow);
        }}
        .title {{ margin: 0; font-size: 28px; letter-spacing: 0.3px; }}
        .subtitle {{ margin: 6px 0 0; opacity: 0.94; font-size: 14px; }}
        .meta {{
            margin-top: 14px;
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
            gap: 10px;
            font-size: 13px;
        }}
        .meta-item {{ background: rgba(255,255,255,0.14); border: 1px solid rgba(255,255,255,0.28); border-radius: 8px; padding: 8px 10px; }}

        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(170px, 1fr));
            gap: 12px;
            margin: 16px 0 18px;
        }}
        .card {{
            background: var(--surface);
            border: 1px solid var(--line);
            border-radius: 12px;
            padding: 14px;
            box-shadow: var(--shadow);
        }}
        .card h3 {{ margin: 0; font-size: 12px; text-transform: uppercase; letter-spacing: 0.5px; color: var(--muted); }}
        .card p {{ margin: 8px 0 0; font-size: 28px; font-weight: 700; color: var(--ink); }}

        .section-title-row {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 10px;
            margin: 24px 0 10px;
            flex-wrap: wrap;
        }}
        .section-title-row .section-title {{ margin: 0; }}
        .analysis-section.analysis-hidden .analysis-grid {{ display: none; }}
        .analysis-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
            gap: 12px;
            margin: 0 0 18px;
        }}
        .analysis-card {{
            background: var(--surface);
            border: 1px solid var(--line);
            border-radius: 12px;
            box-shadow: var(--shadow);
            overflow: hidden;
        }}
        .analysis-card h3 {{
            margin: 0;
            padding: 10px 12px;
            font-size: 13px;
            color: #0b3f66;
            background: #e7f0fb;
            border-bottom: 1px solid #c7d7ec;
        }}
        .analysis-card ul {{ list-style: none; margin: 0; padding: 8px 10px; display: grid; gap: 7px; }}
        .analysis-card li {{
            display: grid;
            grid-template-columns: 1fr auto;
            gap: 8px;
            align-items: start;
            padding: 7px 0;
            border-bottom: 1px solid #edf2f7;
            font-size: 12px;
        }}
        .analysis-card li:last-child {{ border-bottom: none; }}
        .analysis-card strong {{ color: #7f1d1d; white-space: nowrap; }}
        .analysis-empty {{ color: var(--muted); }}
        .analysis-action {{
            border: none;
            background: transparent;
            color: inherit;
            padding: 0;
            text-align: left;
            cursor: pointer;
            font: inherit;
        }}
        .analysis-action:hover span {{ color: #0a4d86; text-decoration: underline; }}

        .section-title {{ margin: 24px 0 10px; font-size: 20px; color: #0b3f66; }}
        .matrix-meta {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 10px;
            flex-wrap: wrap;
            margin: 4px 0 10px;
            color: var(--muted);
            font-size: 12px;
        }}
        .filter-summary {{ font-weight: 700; color: #334155; }}
        .table-wrap {{
            background: var(--surface);
            border: 1px solid var(--line);
            border-radius: 12px;
            overflow: auto;
            box-shadow: var(--shadow);
        }}
        table {{ width: 100%; border-collapse: separate; border-spacing: 0; min-width: 1420px; }}
        th, td {{ border-bottom: 1px solid var(--line); padding: 11px 12px; text-align: left; vertical-align: top; font-size: 13px; }}
        th {{
            position: sticky;
            top: 0;
            z-index: 2;
            background: var(--surface-2);
            color: #0b3f66;
            font-size: 12px;
            text-transform: uppercase;
            letter-spacing: 0.4px;
            border-bottom: 2px solid #c7d7ec;
        }}
        tr:hover td {{ background: #f8fbff; }}
        code {{
            font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
            background: #eef4ff;
            border: 1px solid #d6e4ff;
            border-radius: 6px;
            padding: 2px 6px;
            color: #1d3557;
            font-size: 12px;
        }}
        .status-chip {{ display: inline-block; border-radius: 999px; padding: 2px 10px; font-size: 12px; font-weight: 700; letter-spacing: 0.2px; }}
        .status-pass {{ background: var(--pass-bg); color: var(--pass-fg); border: 1px solid #86efac; }}
        .status-fail {{ background: var(--fail-bg); color: var(--fail-fg); border: 1px solid #fca5a5; }}
        .type-chip {{ display: inline-block; border-radius: 999px; padding: 2px 10px; font-size: 12px; font-weight: 700; letter-spacing: 0.2px; }}
        .test-positive {{ background: #dbeafe; color: #1e3a8a; border: 1px solid #93c5fd; }}
        .test-negative {{ background: #fed7aa; color: #92400e; border: 1px solid #fdba74; }}
        .rt-chip {{ display: inline-block; min-width: 64px; text-align: center; border-radius: 8px; padding: 3px 8px; font-weight: 700; }}
        .rt-fast {{ background: var(--fast-bg); color: var(--fast-fg); }}
        .rt-medium {{ background: var(--mid-bg); color: var(--mid-fg); }}
        .rt-slow {{ background: var(--slow-bg); color: var(--slow-fg); }}
        .reason-chip {{ display: inline-block; max-width: 320px; line-height: 1.35; }}
        .reason-fail {{ color: #7f1d1d; font-weight: 700; }}
        .reason-muted {{ color: #64748b; }}
        .detail-link {{
            color: #0a4d86;
            text-decoration: none;
            font-weight: 600;
            border-bottom: 1px dotted #0a4d86;
        }}
        .detail-link:hover {{ color: #083a64; border-bottom-color: #083a64; }}
        .actual-btn {{
            border: 1px solid #a7c2e8;
            background: #eef4ff;
            color: #0a4d86;
            border-radius: 8px;
            padding: 4px 10px;
            font-weight: 600;
            cursor: pointer;
            font-size: 12px;
        }}
        .actual-btn:hover {{
            background: #dbeafe;
            border-color: #8ab0e0;
        }}

        .modal-overlay {{
            position: fixed;
            inset: 0;
            background: rgba(15, 23, 42, 0.65);
            display: none;
            align-items: center;
            justify-content: center;
            z-index: 1000;
            padding: 18px;
        }}
        .modal-overlay.open {{ display: flex; }}
        .modal-card {{
            width: min(980px, 96vw);
            max-height: 88vh;
            background: #ffffff;
            border-radius: 12px;
            box-shadow: 0 14px 42px rgba(15, 23, 42, 0.38);
            border: 1px solid #d8e5f4;
            overflow: hidden;
            display: flex;
            flex-direction: column;
        }}
        .modal-head {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            padding: 12px 14px;
            background: #e7f0fb;
            border-bottom: 1px solid #c7d7ec;
        }}
        .modal-head h3 {{ margin: 0; font-size: 15px; color: #0b3f66; }}
        .modal-close {{
            border: none;
            background: transparent;
            color: #0b3f66;
            font-size: 24px;
            line-height: 1;
            cursor: pointer;
            padding: 0 4px;
        }}
        .modal-meta {{
            display: flex;
            flex-wrap: wrap;
            gap: 14px;
            padding: 8px 14px;
            font-size: 12px;
            color: #475569;
            border-bottom: 1px solid #e2e8f0;
            background: #f8fbff;
        }}
        .modal-pre {{
            margin: 0;
            padding: 14px;
            overflow: auto;
            white-space: pre-wrap;
            color: #dbeafe;
            background: linear-gradient(180deg, #0c1930 0%, #111f38 100%);
            font-size: 12px;
            line-height: 1.55;
            flex: 1;
        }}

        .details-wrap {{ margin-top: 14px; display: grid; gap: 14px; }}
        .details-wrap.details-hidden {{ display: none; }}
        .detail-card {{
            background: var(--surface);
            border: 1px solid var(--line);
            border-radius: 12px;
            box-shadow: var(--shadow);
            padding: 14px;
            scroll-margin-top: 24px;
        }}
        .detail-head h3 {{ margin: 0; font-size: 16px; color: #0b3f66; }}
        .detail-meta {{ margin-top: 4px; color: var(--muted); font-size: 12px; }}
        .json-grid {{
            margin-top: 10px;
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 10px;
        }}
        .span-2 {{ grid-column: 1 / -1; }}
        article {{
            border: 1px solid #d8e5f4;
            border-radius: 10px;
            overflow: hidden;
            background: #f8fbff;
        }}
        article h4 {{
            margin: 0;
            padding: 8px 10px;
            font-size: 12px;
            text-transform: uppercase;
            letter-spacing: 0.4px;
            color: #0b3f66;
            background: #e7f0fb;
            border-bottom: 1px solid #c7d7ec;
        }}
        pre {{
            margin: 0;
            padding: 10px;
            white-space: pre-wrap;
            overflow-x: auto;
            max-height: 300px;
            font-size: 12px;
            line-height: 1.5;
            color: #dbeafe;
            background: linear-gradient(180deg, #0c1930 0%, #111f38 100%);
        }}
        .footer {{ margin: 16px 2px 4px; color: #516076; font-size: 12px; }}

        @media (max-width: 900px) {{
            .container {{ padding: 14px; }}
            .layout {{ grid-template-columns: 1fr; }}
            .service-sidebar {{ position: static; max-height: none; }}
            .toolbar {{ justify-content: flex-start; }}
            .title {{ font-size: 23px; }}
            .json-grid {{ grid-template-columns: 1fr; }}
            .span-2 {{ grid-column: auto; }}
        }}
        @media print {{
            body {{ background: #ffffff; color: #111827; }}
            .container {{ max-width: none; padding: 0; }}
            .layout {{ display: block; }}
            .toolbar,
            .service-sidebar,
            .modal-overlay,
            #back-to-top,
            #details-title-row,
            .details-wrap,
            .footer {{ display: none !important; }}
            .header,
            .card,
            .analysis-card,
            .table-wrap {{ box-shadow: none; }}
            .header {{ color: #111827; background: #ffffff; border: 1px solid #d1d5db; }}
            .summary,
            .analysis-grid {{ break-inside: avoid; }}
            table {{ min-width: 0; font-size: 10px; }}
            th, td {{ padding: 6px; font-size: 10px; }}
            .analysis-section.analysis-hidden .analysis-grid {{ display: grid; }}
        }}
        /* Back-to-top floating button */
        #back-to-top {{
            position: fixed;
            right: 20px;
            bottom: 20px;
            width: 44px;
            height: 44px;
            border-radius: 50%;
            background: var(--brand);
            color: #fff;
            border: none;
            display: none;
            align-items: center;
            justify-content: center;
            font-size: 22px;
            cursor: pointer;
            box-shadow: 0 8px 20px rgba(15,23,42,0.18);
            z-index: 1200;
        }}
        #back-to-top.visible {{ display: flex; opacity: 1; transform: translateY(0); transition: opacity 200ms, transform 200ms; }}
        #back-to-top.hidden {{ opacity: 0; transform: translateY(8px); transition: opacity 200ms, transform 200ms; }}
  </style>
</head>
<body>
    <main class="container">
        <div class="layout">
            {sidebar_html}
            <section class="content">
        <div class="toolbar">
            <div class="filter-panel">
                <div class="status-filters">
                    <button type="button" class="filter-btn active" data-status-filter="all">All</button>
                    <button type="button" class="filter-btn pass-btn" data-status-filter="passed">&#10003; Passed</button>
                    <button type="button" class="filter-btn fail-btn" data-status-filter="failed">&#10007; Failed</button>
                </div>
                <select id="api-filter" class="filter-select" aria-label="Filter by API">
                    <option value="">All APIs</option>
                    {api_options}
                </select>
                <select id="test-type-filter" class="filter-select" aria-label="Filter by test type">
                    <option value="">All Test Types</option>
                    {test_type_options}
                </select>
                <input id="search-filter" class="search-input" type="search" placeholder="Search cases" aria-label="Search cases" />
            </div>
            <div class="toolbar-actions">
                <button type="button" id="reset-filters" class="utility-btn">Reset Filters</button>
                <button type="button" id="analysis-toggle" class="utility-btn" aria-expanded="true">Hide Analysis</button>
                <button type="button" id="details-toggle" class="utility-btn" aria-expanded="true">Hide Details</button>
                <button type="button" id="sidebar-toggle" class="sidebar-toggle" aria-expanded="true">Hide Services</button>
            </div>
        </div>
        <section class="header">
            <h1 class="title">SIT Automation Execution Report</h1>
            <p class="subtitle">Banking payment switch validation summary and technical traceability</p>
            <div class="meta">
                <div class="meta-item"><strong>Run ID:</strong> {run_tag}</div>
                <div class="meta-item"><strong>Started:</strong> {started_at.isoformat()}</div>
                <div class="meta-item"><strong>Finished:</strong> {finished_at.isoformat()}</div>
                <div class="meta-item"><strong>Duration:</strong> {duration_seconds:.2f} seconds</div>
            </div>
        </section>

        <section class="summary">
            <article class="card"><h3>Total Test Cases</h3><p>{summary['total']}</p></article>
            <article class="card"><h3>Passed</h3><p>{summary['passed']}</p></article>
            <article class="card"><h3>Failed</h3><p>{summary['failed']}</p></article>
            <article class="card"><h3>Pass Rate</h3><p>{pass_rate}%</p></article>
            <article class="card"><h3>Avg Response Time</h3><p>{average_response_ms} ms</p></article>
        </section>

        <section id="failure-analysis-section" class="{analysis_section_class}" data-default-hidden="{initial_analysis_hidden}">
            <div class="section-title-row">
                <h2 class="section-title">Failure Analysis</h2>
            </div>
            {failure_analysis_html}
        </section>

        <h2 class="section-title">Execution Matrix</h2>
        <div class="matrix-meta">
            <span id="filter-summary" class="filter-summary">Showing {summary['total']} of {summary['total']} cases</span>
            <span>Use filters, search, or Failure Analysis entries to focus the run.</span>
        </div>
        <div class="table-wrap">
            <table>
                <thead>
                    <tr>
                        <th>TC ID</th>
                        <th>Service</th>
                        <th>Description</th>
                        <th>Test Type</th>
                        <th>Bank</th>
                        <th>API</th>
                        <th>Status</th>
                        <th>Expected</th>
                        <th>Actual</th>
                        <th>Response (ms)</th>
                        <th>Details</th>
                        <th>Failure Reason</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(table_rows)}
                </tbody>
            </table>
        </div>

        <div id="details-title-row" class="section-title-row">
            <h2 class="section-title">API Request and Response Details</h2>
        </div>
        <div id="details-wrap" class="{details_wrap_class}" data-default-hidden="{initial_details_hidden}">
            {''.join(details_sections)}
        </div>

        <p class="footer">Generated by SIT Automation Framework</p>
            </section>
        </div>
    </main>
    <button id="back-to-top" aria-label="Back to top" title="Back to top">&#9650;</button>
    <script>
        (function() {{
            function closeAllModals() {{
                document.querySelectorAll('.modal-overlay.open').forEach(function(modal) {{
                    modal.classList.remove('open');
                    modal.setAttribute('aria-hidden', 'true');
                }});
            }}

            document.querySelectorAll('.actual-btn[data-modal-target]').forEach(function(button) {{
                button.addEventListener('click', function() {{
                    var modalId = button.getAttribute('data-modal-target');
                    var modal = document.getElementById(modalId);
                    if (!modal) {{
                        return;
                    }}
                    closeAllModals();
                    modal.classList.add('open');
                    modal.setAttribute('aria-hidden', 'false');
                }});
            }});

            document.addEventListener('click', function(event) {{
                var target = event.target;
                if (!(target instanceof Element)) {{
                    return;
                }}
                if (target.matches('[data-modal-close="true"]')) {{
                    closeAllModals();
                    return;
                }}
                if (target.classList.contains('modal-overlay')) {{
                    closeAllModals();
                }}
            }});

            document.addEventListener('keydown', function(event) {{
                if (event.key === 'Escape') {{
                    closeAllModals();
                }}
            }});

            var activeStatus = 'all';
            var activeService = null;
            var activeApi = '';
            var activeTestType = '';
            var activeSearch = '';
            var totalRows = document.querySelectorAll('.result-row').length;
            var filterSummary = document.getElementById('filter-summary');
            var analysisStorageKey = 'sit-report-hide-analysis';
            var detailsStorageKey = 'sit-report-hide-details';

            function setStatusFilter(status) {{
                activeStatus = status || 'all';
                document.querySelectorAll('.filter-btn[data-status-filter]').forEach(function(button) {{
                    button.classList.toggle('active', button.getAttribute('data-status-filter') === activeStatus);
                }});
            }}

            function setSearch(value) {{
                activeSearch = (value || '').trim().toLowerCase();
                if (searchFilter) {{
                    searchFilter.value = value || '';
                }}
            }}

            function setApi(value) {{
                activeApi = value || '';
                if (apiFilter) {{
                    apiFilter.value = activeApi;
                }}
            }}

            function setTestType(value) {{
                activeTestType = value || '';
                if (testTypeFilter) {{
                    testTypeFilter.value = activeTestType;
                }}
            }}

            function applyFilters() {{
                var visibleRows = 0;
                document.querySelectorAll('.result-row').forEach(function(row) {{
                    var serviceMatch = !activeService || row.getAttribute('data-service') === activeService;
                    var status = row.getAttribute('data-row-status') || '';
                    var statusMatch = activeStatus === 'all' || status === activeStatus;
                    var apiMatch = !activeApi || row.getAttribute('data-api') === activeApi;
                    var testTypeMatch = !activeTestType || row.getAttribute('data-test-type') === activeTestType;
                    var searchText = row.getAttribute('data-search') || '';
                    var searchMatch = !activeSearch || searchText.indexOf(activeSearch) !== -1;
                    var visible = serviceMatch && statusMatch && apiMatch && testTypeMatch && searchMatch;
                    row.style.display = visible ? '' : 'none';
                    if (visible) {{
                        visibleRows += 1;
                    }}
                }});
                document.querySelectorAll('.result-detail').forEach(function(card) {{
                    var serviceMatch = !activeService || card.getAttribute('data-service') === activeService;
                    var status = card.getAttribute('data-detail-status') || '';
                    var statusMatch = activeStatus === 'all' || status === activeStatus;
                    var apiMatch = !activeApi || card.getAttribute('data-api') === activeApi;
                    var testTypeMatch = !activeTestType || card.getAttribute('data-test-type') === activeTestType;
                    var searchText = card.getAttribute('data-search') || '';
                    var searchMatch = !activeSearch || searchText.indexOf(activeSearch) !== -1;
                    card.style.display = serviceMatch && statusMatch && apiMatch && testTypeMatch && searchMatch ? '' : 'none';
                }});
                if (filterSummary) {{
                    filterSummary.textContent = 'Showing ' + visibleRows + ' of ' + totalRows + ' cases';
                }}
            }}

            function filterByService(serviceKey) {{
                activeService = serviceKey;
                document.querySelectorAll('.service-link').forEach(function(button) {{
                    var buttonService = button.getAttribute('data-service-filter') || null;
                    button.classList.toggle('active', buttonService === serviceKey);
                }});
                applyFilters();
            }}

            var serviceButtons = Array.prototype.slice.call(document.querySelectorAll('.service-link[data-service-filter]'));
            serviceButtons.forEach(function(button) {{
                button.addEventListener('click', function() {{
                    var serviceKey = button.getAttribute('data-service-filter');
                    filterByService(serviceKey || null);
                }});
            }});

            document.querySelectorAll('.filter-btn[data-status-filter]').forEach(function(btn) {{
                btn.addEventListener('click', function() {{
                    setStatusFilter(btn.getAttribute('data-status-filter'));
                    applyFilters();
                }});
            }});

            var apiFilter = document.getElementById('api-filter');
            if (apiFilter) {{
                apiFilter.addEventListener('change', function() {{
                    activeApi = apiFilter.value || '';
                    applyFilters();
                }});
            }}

            var testTypeFilter = document.getElementById('test-type-filter');
            if (testTypeFilter) {{
                testTypeFilter.addEventListener('change', function() {{
                    activeTestType = testTypeFilter.value || '';
                    applyFilters();
                }});
            }}

            var searchFilter = document.getElementById('search-filter');
            if (searchFilter) {{
                searchFilter.addEventListener('input', function() {{
                    setSearch(searchFilter.value || '');
                    applyFilters();
                }});
            }}

            var resetFilters = document.getElementById('reset-filters');
            if (resetFilters) {{
                resetFilters.addEventListener('click', function() {{
                    filterByService(null);
                    setStatusFilter('all');
                    setApi('');
                    setTestType('');
                    setSearch('');
                    applyFilters();
                }});
            }}

            document.querySelectorAll('.analysis-action[data-analysis-action]').forEach(function(button) {{
                button.addEventListener('click', function() {{
                    var action = button.getAttribute('data-analysis-action') || '';
                    var value = button.getAttribute('data-analysis-value') || '';
                    setStatusFilter('failed');
                    if (action === 'api') {{
                        setApi(value);
                        setSearch('');
                    }} else {{
                        setSearch(value);
                    }}
                    applyFilters();
                    var matrix = document.querySelector('.table-wrap');
                    if (matrix) {{
                        matrix.scrollIntoView({{ behavior: 'smooth', block: 'start' }});
                    }}
                }});
            }});

            function readStoredHidden(key, fallback) {{
                try {{
                    var value = window.localStorage.getItem(key);
                    if (value === 'true') {{
                        return true;
                    }}
                    if (value === 'false') {{
                        return false;
                    }}
                }} catch (error) {{
                    return fallback;
                }}
                return fallback;
            }}

            function writeStoredHidden(key, hidden) {{
                try {{
                    window.localStorage.setItem(key, hidden ? 'true' : 'false');
                }} catch (error) {{
                    return;
                }}
            }}

            var analysisSection = document.getElementById('failure-analysis-section');
            var analysisToggle = document.getElementById('analysis-toggle');
            function setAnalysisHidden(hidden, persist) {{
                if (!analysisSection || !analysisToggle) {{
                    return;
                }}
                analysisSection.classList.toggle('analysis-hidden', hidden);
                analysisToggle.textContent = hidden ? 'Show Analysis' : 'Hide Analysis';
                analysisToggle.setAttribute('aria-expanded', hidden ? 'false' : 'true');
                if (persist) {{
                    writeStoredHidden(analysisStorageKey, hidden);
                }}
            }}
            if (analysisSection && analysisToggle) {{
                var defaultAnalysisHidden = analysisSection.getAttribute('data-default-hidden') === 'true';
                setAnalysisHidden(readStoredHidden(analysisStorageKey, defaultAnalysisHidden), false);
                analysisToggle.addEventListener('click', function() {{
                    setAnalysisHidden(!analysisSection.classList.contains('analysis-hidden'), true);
                }});
            }}

            var detailsWrap = document.getElementById('details-wrap');
            var detailsToggle = document.getElementById('details-toggle');
            function setDetailsHidden(hidden, persist) {{
                if (!detailsWrap || !detailsToggle) {{
                    return;
                }}
                detailsWrap.classList.toggle('details-hidden', hidden);
                detailsToggle.textContent = hidden ? 'Show Details' : 'Hide Details';
                detailsToggle.setAttribute('aria-expanded', hidden ? 'false' : 'true');
                if (persist) {{
                    writeStoredHidden(detailsStorageKey, hidden);
                }}
            }}
            if (detailsWrap && detailsToggle) {{
                var defaultDetailsHidden = detailsWrap.getAttribute('data-default-hidden') === 'true';
                setDetailsHidden(readStoredHidden(detailsStorageKey, defaultDetailsHidden), false);
                detailsToggle.addEventListener('click', function() {{
                    setDetailsHidden(!detailsWrap.classList.contains('details-hidden'), true);
                }});
            }}

            var layout = document.querySelector('.layout');
            var sidebarToggle = document.getElementById('sidebar-toggle');
            if (layout && sidebarToggle) {{
                sidebarToggle.addEventListener('click', function() {{
                    var collapsed = layout.classList.toggle('sidebar-collapsed');
                    sidebarToggle.textContent = collapsed ? 'Show Services' : 'Hide Services';
                    sidebarToggle.setAttribute('aria-expanded', collapsed ? 'false' : 'true');
                }});
            }}

            // Back-to-top button behavior
            var backToTop = document.getElementById('back-to-top');
            if (backToTop) {{
                function updateBackToTop() {{
                    if (window.scrollY > 300) {{
                        backToTop.classList.add('visible');
                        backToTop.classList.remove('hidden');
                    }} else {{
                        backToTop.classList.remove('visible');
                        backToTop.classList.add('hidden');
                    }}
                }}

                backToTop.addEventListener('click', function() {{
                    window.scrollTo({{ top: 0, behavior: 'smooth' }});
                }});

                backToTop.addEventListener('keydown', function(e) {{
                    if (e.key === 'Enter' || e.key === ' ') {{
                        e.preventDefault();
                        backToTop.click();
                    }}
                }});

                window.addEventListener('scroll', updateBackToTop);
                // initialize
                updateBackToTop();
            }}

            applyFilters();
        }})();
    </script>
</body>
</html>
"""
        report_file.write_text(html, encoding="utf-8")

    @staticmethod
    def _format_actual_for_modal(actual_text: str) -> str:
        stripped = actual_text.strip()
        if not stripped:
            return ""

        try:
            parsed = json.loads(stripped)
            return json.dumps(parsed, indent=2, ensure_ascii=False)
        except json.JSONDecodeError:
            return actual_text

    @staticmethod
    def _safe_csv(value: Any) -> str:
        text = str(value if value is not None else "")
        text = text.replace('"', '""')
        return f'"{text}"'

    @staticmethod
    def _build_summary(results: List[Dict[str, Any]]) -> Dict[str, Any]:
        total = len(results)
        passed = sum(1 for result in results if result.get("status") == "PASSED")
        failed = total - passed
        pass_percent = round((passed / total) * 100, 2) if total else 0.0
        return {
            "total": total,
            "passed": passed,
            "failed": failed,
            "pass_percent": pass_percent,
        }

    @staticmethod
    def _html_escape(text: str) -> str:
        return (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
            .replace("'", "&#39;")
        )


if __name__ == "__main__":
    root = Path(__file__).resolve().parent
    runner = SITRunner(root)
    runner.run()
